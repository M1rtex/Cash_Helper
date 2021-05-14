import sys
import platform
from PyQt5 import QtCore
from PySide2 import QtCore, QtGui, QtWidgets
from PySide2.QtCore import (QCoreApplication, Qt, QPropertyAnimation)
from PySide2.QtWidgets import *
from win32api import GetSystemMetrics
import openpyxl
import time

import icons
import main_iu
import starting

WINDOW_SIZE = 0
class Starting(QMainWindow, starting.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        self.counter = 0

        self.timer = QtCore.QTimer()
        self.timer.timeout.connect(self.progress)
        self.timer.start(35)

        self.timer.singleShot(500, lambda: self.descriptionLable.setText('<strong>Welcome</strong> to Application'))
        self.timer.singleShot(2000, lambda: self.descriptionLable.setText('<strong>LOADING</strong> Databases'))
        self.timer.singleShot(3000, lambda: self.descriptionLable.setText('<strong>LOADING</strong> User Interface'))

    def progress(self):
        self.progressBar.setValue(self.counter)
        if self.counter >= 100:
            self.timer.stop()
            main()
            self.close()
        self.counter += 1



class Theapp(QMainWindow, main_iu.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.start = True
        self.currency = 'RUB'
        self.pleasure_percent = 0
        self.rest_percent = 0
        self.remains_percent = 0
        self.setupUi(self)  # Это нужно для инициализации дизайна
        self.minimizeButton.clicked.connect(lambda: self.showMinimized())
        self.restoreButton.clicked.connect(lambda: self.restore_or_minimize())
        self.closeButton.clicked.connect(lambda: self.close())
        self.left_men_togle_btn.clicked.connect(lambda: self.slideLeftMenu())
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        self.stackedWidget.setCurrentWidget(self.home_page)
        self.homeButton.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.home_page))
        self.statisticButton.clicked.connect(lambda: self.gotoStatistic())
        self.MoreInfoButton.clicked.connect(lambda: self.barssetup())
        self.BackInfoButton.clicked.connect(lambda: self.barsdisable())
        self.settingsButton.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.settings_page))
        self.ConfirmButton.clicked.connect(lambda: self.info_processing())
        self.SettingsConfirmButton.clicked.connect(lambda: self.get_curency())
        self.ClearLOGButton.clicked.connect(lambda: self.clean_LOG())
        self.SaveAgainButton.clicked.connect(lambda: self.save_again())
        self.Log_lable.setText('App is running.')
        self.PleasureBar.setValue(0)
        self.RestBar.setValue(0)
        self.RemainsBar.setValue(0)
        self.progressBarValue(self.info_collection())
        def moveWindow(e):
            if self.isMaximized() == False:
                if e.buttons() == Qt.LeftButton:
                    self.move(self.pos() + e.globalPos() - self.clickPosition)
                    self.clickPosition = e.globalPos()
                    e.accept()
        self.main_header.mouseMoveEvent = moveWindow
    def mousePressEvent(self, event):
        self.clickPosition = event.globalPos()
    def restore_or_minimize(self):
        global WINDOW_SIZE
        width = GetSystemMetrics(0) - 150
        height = GetSystemMetrics(1) - 150
        if WINDOW_SIZE == 0:
            self.SizeAnimation = QPropertyAnimation(self, b'size')
            self.SizeAnimation.setEndValue(QtCore.QSize(width,height))
            self.SizeAnimation.setDuration(2000)
            self.SizeAnimation.setEasingCurve(QtCore.QEasingCurve.InOutQuart)
            self.SizeAnimation.start()
            WINDOW_SIZE = 1
        else:
            self.SizeAnimation = QPropertyAnimation(self, b'size')
            self.SizeAnimation.setEndValue(QtCore.QSize(500,500))
            self.SizeAnimation.setDuration(2000)
            self.SizeAnimation.setEasingCurve(QtCore.QEasingCurve.InOutQuart)
            self.SizeAnimation.start()
            WINDOW_SIZE = 0
    def slideLeftMenu(self):
        width = self.left_menu.width()
        if width == 45:
            newWidth = 90
            self.slideanim = QPropertyAnimation(self.left_menu, b'maximumWidth')
            self.slideanim.setDuration(100)
            self.slideanim.setStartValue(45)
            self.slideanim.setEndValue(newWidth)
            self.slideanim.setEasingCurve(QtCore.QEasingCurve.InQuad)
            self.slideanim.start()
            self.homeButton.setText('HOME')
            self.homeButton.setMinimumWidth(90)
            self.statisticButton.setText('STATISTIC')
            self.statisticButton.setMinimumWidth(90)
            self.settingsButton.setText('SETTINGS')
            self.settingsButton.setMinimumWidth(90)
        else:
            newWidth = 45
            self.slideanim = QPropertyAnimation(self.left_menu, b'maximumWidth')
            self.slideanim.setDuration(100)
            self.slideanim.setStartValue(90)
            self.slideanim.setEndValue(newWidth)
            self.slideanim.setEasingCurve(QtCore.QEasingCurve.InQuad)
            self.slideanim.start()
            self.homeButton.setText('')
            self.homeButton.setMinimumWidth(45)
            self.statisticButton.setText('')
            self.statisticButton.setMinimumWidth(45)
            self.settingsButton.setText('')
            self.settingsButton.setMinimumWidth(45)
        # self.left_menu.setMaximumWidth(newWidth)
    def gotoStatistic(self):
        self.progressBarValue(self.info_collection())
        self.stackedWidget.setCurrentWidget(self.Statistic_page)
    def info_processing(self):
        self.salary = self.Salary_lineEdit.text()
        self.pleasure = self.Pleasure_lineEdit.text()
        self.rest = self.Rest_lineEdit.text()
        if self.salary == '':
            res = 'The SALARY was entered incorrectly.'
            self.Log_lable.setText(res)
        elif self.rest == '':
            res = 'The REST was entered incorrectly.'
            self.Log_lable.setText(res)
        elif self.pleasure == '':
            res = 'The PlEASURE was entered incorrectly.'
            self.Log_lable.setText(res)
        else:
            res = 'Saving...\n'
            self.Log_lable.setText(res)
            res += 'Saved!'
            self.salary = int(self.salary)
            self.rest = int(self.rest)
            self.pleasure = int(self.pleasure)
            self.pleasure_percent = (self.pleasure / self.salary) * 100
            self.rest_percent = (self.rest / self.salary) * 100
            remains = self.salary - (self.pleasure + self.rest)
            self.remains_percent = (remains / self.salary) * 100
            try:
                wb = openpyxl.load_workbook('LOG.xlsx')
            except FileNotFoundError:
                wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'LOG'
            ws['A1'] = 'Дата, время'
            ws['B1'] = 'Зарплата'
            ws['C1'] = 'Развлечения, % от зарплаты'
            ws['D1'] = 'Остальное, % от зарплаты'
            ws['E1'] = 'Остаток, % от зарплаты'
            ws['F1'] = 'Валюта'
            self.date = time.strftime("%H:%M %B %d, %Y", time.localtime())
            i = 2
            self.specified = 30
            while i != self.specified:
                if ws[f'A{i}'].value == None:
                    ws[f'A{i}'] = self.date
                    ws[f'B{i}'] = self.salary
                    ws[f'C{i}'] = f'{round(self.pleasure_percent, 2)}%: {self.pleasure}'
                    ws[f'D{i}'] = f'{round(self.rest_percent, 2)}%: {self.rest}'
                    ws[f'E{i}'] = f'{round(self.remains_percent, 2)}%: {remains}'
                    ws[f'F{i}'] = self.currency
                    break
                i += 1
            if self.start:
                self.save_path = QFileDialog.getSaveFileName()[0]
                self.start = False
            wb.save(self.save_path + '.xlsx')
            wb.save('LOG.xlsx')
            self.Log_lable.setText(res)
    def info_collection(self):
        try:
            wb = openpyxl.load_workbook('LOG.xlsx')
            ws = wb.active
            item = 2
            sum = 0
            salary = 0
            while ws[f'A{item}'].value != None:
                salar = ws[f'B{item}'].value
                pleasure = ws[f'C{item}'].value
                pleasure = pleasure.split(' ')
                rest = ws[f'D{item}'].value
                rest = rest.split(' ')
                sum += int(pleasure[1]) + int(rest[1])
                salary += int(round(salar))
                item += 1
            try:
                percent = (sum / salary) * 100
            except ZeroDivisionError:
                percent = 0
            return int(round(percent))
        except FileNotFoundError:
            percent = 0
            return percent
    def progressBarValue(self, value):
        styleSheet = '''
        QFrame {
            border-radius: 150;
            background-color: qconicalgradient(cx:0.5, cy:0.5, angle:90, stop:{STOP_1} rgba(34, 0, 146, 0), stop:{STOP_2} rgba(158, 158, 158, 255));
        }
        '''
        progress = (100 - value) / 100.0
        stop_1 = str(progress - 0.001)
        stop_2 = str(progress)
        newStyleSheet = styleSheet.replace('{STOP_1}', stop_1).replace('{STOP_2}', stop_2)
        self.CircularProgress.setStyleSheet(newStyleSheet)
        self.coastsPercent.setText('{value}%'.replace('{value}', str(value)))

    def barsValue(self, pleasureValue, restValue):
        if pleasureValue != 0 or restValue != 0:
            self.pleasureanim = QPropertyAnimation(self.PleasureBar, b'value')
            self.pleasureanim.setDuration(2000)
            self.pleasureanim.setStartValue(0)
            self.pleasureanim.setEndValue(pleasureValue)
            self.pleasureanim.setEasingCurve(QtCore.QEasingCurve.Linear)
            self.pleasureanim.start()
            if pleasureValue >= 45:
                self.PleasureBar.setStyleSheet('''
                        QProgressBar{
                            border-radius: 5;
                            background-color: rgb(118, 118, 118);
                            color: rgb(0,0,0);
                        }
        
                        QProgressBar::chunk{
                            border-radius: 5;
                            background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(133, 134, 146, 255), stop:1 rgba(255, 255, 255, 255));
                        }
                        ''')
            else:
                self.PleasureBar.setStyleSheet('''
                QProgressBar{
                    border-radius: 5;
                    background-color: rgb(118, 118, 118);
                    color: white;
                }
    
                QProgressBar::chunk{
                    border-radius: 5;
                    background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(133, 134, 146, 255), stop:1 rgba(255, 255, 255, 255));
                }
                ''')

            self.restanim = QPropertyAnimation(self.RestBar, b'value')
            self.restanim.setDuration(950)
            self.restanim.setStartValue(0)
            self.restanim.setEndValue(restValue)
            self.restanim.setEasingCurve(QtCore.QEasingCurve.Linear)
            self.restanim.start()
            if restValue >= 45:
                self.RestBar.setStyleSheet('''
                        QProgressBar{
                            border-radius: 5;
                            background-color: rgb(118, 118, 118);
                            color: rgb(0,0,0);
                        }
        
                        QProgressBar::chunk{
                            border-radius: 5;
                            background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(133, 134, 146, 255), stop:1 rgba(255, 255, 255, 255));
                        }
                        ''')
            else:
                self.RestBar.setStyleSheet('''
                QProgressBar{
                    border-radius: 5;
                    background-color: rgb(118, 118, 118);
                    color: white;
                }
    
                QProgressBar::chunk{
                    border-radius: 5;
                    background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(133, 134, 146, 255), stop:1 rgba(255, 255, 255, 255));
                }
                ''')

            remainsValue = 100 - (pleasureValue+restValue)
            self.remainsAnim = QPropertyAnimation(self.RemainsBar, b'value')
            self.remainsAnim.setDuration(950)
            self.remainsAnim.setStartValue(0)
            self.remainsAnim.setEndValue(remainsValue)
            self.remainsAnim.setEasingCurve(QtCore.QEasingCurve.Linear)
            self.remainsAnim.start()
            if remainsValue >= 45:
                self.RemainsBar.setStyleSheet('''
                QProgressBar{
                    border-radius: 5;
                    background-color: rgb(118, 118, 118);
                    color: rgb(0,0,0);
                }
    
                QProgressBar::chunk{
                    border-radius: 5;
                    background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(133, 134, 146, 255), stop:1 rgba(255, 255, 255, 255));
                }
                ''')
            else:
                self.RemainsBar.setStyleSheet('''
                QProgressBar{
                    border-radius: 5;
                    background-color: rgb(118, 118, 118);
                    color: white;
                }
    
                QProgressBar::chunk{
                    border-radius: 5;
                    background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(133, 134, 146, 255), stop:1 rgba(255, 255, 255, 255));
                }
                ''')
    def barssetup(self):
        self.stackedWidget.setCurrentWidget(self.MoreStatistic_page)
        self.barsValue(self.pleasure_percent, self.rest_percent)
    def barsdisable(self):
        self.PleasureBar.setValue(0)
        self.RestBar.setValue(0)
        self.RemainsBar.setValue(0)
        self.stackedWidget.setCurrentWidget(self.Statistic_page)
    def get_curency(self):
        self.currency = self.CurrencyBox.currentText()
    def clean_LOG(self):
        try:
            wb = openpyxl.load_workbook('LOG.xlsx')
            ws = wb.active
            item = 2
            while ws[f'A{item}'].value != None:
                ws[f'A{item}'] = None
                ws[f'B{item}'] = None
                ws[f'C{item}'] = None
                ws[f'D{item}'] = None
                ws[f'E{item}'] = None
                ws[f'F{item}'] = None
                item += 1
            self.pleasure_percent = 0
            self.rest_percent = 0
            self.remains_percent = 0
            wb.save('LOG.xlsx')
        except FileNotFoundError:
            pass
    def save_again(self):
        wb = openpyxl.load_workbook('LOG.xlsx')
        save_path = QFileDialog.getSaveFileName()[0]
        wb.save(save_path + '.xlsx')
app = QApplication(sys.argv)
Window = Theapp()
Start_Window = Starting()
def start():
    Start_Window.show()
    app.exec_()
def main():
    Window.show()
    app.exec_()
start()